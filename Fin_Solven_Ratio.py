import pandas as pd
import numpy as np
import openpyxl
import xlsxwriter
import subprocess

workbook = openpyxl.load_workbook("Financial_Statement.xlsx")
sheet1 = workbook['balance sheet 2018-2022']
sheet2 = workbook['P&L 2018-2022']
sheet3 = workbook['balance sheet 2013-2017']
sheet4 = workbook['P&L 2013-2017']

def main():
    array1 = np.array(Debt_Equity_Ratio())
    array2 = np.array(Solvency_Ratio())
    array3 = np.array(Propreitary_Ratio())
    array4 = np.array(FixedAsset_Ratio())

    wb = xlsxwriter.Workbook(f'Financial_Ratio.xlsx')
    ws = wb.add_worksheet('Solvency Ratios')

    wb.close()

    Header = ['Mar 22', 'Mar 21', 'Mar 20', 'Mar 19', 'Mar 18', 'Mar 17', 'Mar 16', 'Mar 15', 'Mar 14', 'Mar 13']
    df = pd.DataFrame ({' ' : Header,
                        'Debt Equity Ratio': array1,
                        'Solvency Ratio': array2,
                        'Propreitary Ratio': array3,
                        'Fixed Asset Ratio': array4})
    with pd.ExcelWriter(f'Financial_Ratio.xlsx', mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name = 'Solvency Ratios')

    subprocess.run('Python3 Fin_Liquidity_Ratio.py', shell=True)
    subprocess.run('Python3 Fin_Profit_Ratio.py', shell=True)




def Debt_Equity_Ratio():
    TNCD = []
    TCD = []
    Equi_Shareholder = []
    Total_Non_Current_Debt = []
    Total_Current_Debt = []
    Shareholders_Equity = []
    Debt_Equity_Ratio = []

    rowS1TNCD = sheet1.iter_cols(min_row=12, max_row=12, min_col=2, max_col=6)
    rowS1TCD = sheet1.iter_cols(min_row=18, max_row=18, min_col=2, max_col=6)
    rowS1ES = sheet1.iter_cols(min_row=10, max_row=10, min_col=2, max_col=6)
    for a in rowS1TNCD:
        TNCD.append(a[0].value)
        Total_Non_Current_Debt= np.asfarray(TNCD)
    for a in rowS1TCD:
        TCD.append(a[0].value)
        Total_Current_Debt= np.asfarray(TCD)

    for a in rowS1ES:
        Equi_Shareholder.append(a[0].value)
        Shareholders_Equity = np.asfarray(Equi_Shareholder)

    rowS3TNCD = sheet3.iter_cols(min_row=12, max_row=12, min_col=2, max_col=6)
    rowS3TCD = sheet3.iter_cols(min_row=18, max_row=18, min_col=2, max_col=6)
    rowS3ES = sheet3.iter_cols(min_row=10, max_row=10, min_col=2, max_col=6)
    for a in rowS3TNCD:
        TNCD.append(a[0].value)
        Total_Non_Current_Debt= np.asfarray(TNCD)
    for a in rowS3TCD:
        TCD.append(a[0].value)
        Total_Current_Debt= np.asfarray(TCD)

    for a in rowS3ES:
        Equi_Shareholder.append(a[0].value)
        Shareholders_Equity = np.asfarray(Equi_Shareholder)

    Debt_Equity_Ratio = (Total_Non_Current_Debt+Total_Current_Debt)/Shareholders_Equity
    return (Debt_Equity_Ratio)

def Solvency_Ratio():
    TNCD = []
    TCD = []
    TangAsst = []
    Total_Non_Current_Debt = []
    Total_Current_Debt = []
    Tangible_Assets = []
    Solvency_Ratio = []

    rowS1TNCD = sheet1.iter_cols(min_row=12, max_row=12, min_col=2, max_col=6)
    rowS1TCD = sheet1.iter_cols(min_row=18, max_row=18, min_col=2, max_col=6)
    rowS1TA = sheet1.iter_cols(min_row=26, max_row=26, min_col=2, max_col=6)
    for a in rowS1TNCD:
        TNCD.append(a[0].value)
        Total_Non_Current_Debt= np.asfarray(TNCD)
    for a in rowS1TCD:
        TCD.append(a[0].value)
        Total_Current_Debt= np.asfarray(TCD)

    for a in rowS1TA:
        TangAsst.append(a[0].value)
        Tangible_Assets = np.asfarray(TangAsst)

    rowS3TNCD = sheet3.iter_cols(min_row=12, max_row=12, min_col=2, max_col=6)
    rowS3TCD = sheet3.iter_cols(min_row=18, max_row=18, min_col=2, max_col=6)
    rowS3TA = sheet3.iter_cols(min_row=26, max_row=26, min_col=2, max_col=6)
    for a in rowS3TNCD:
        TNCD.append(a[0].value)
        Total_Non_Current_Debt= np.asfarray(TNCD)
    for a in rowS3TCD:
        TCD.append(a[0].value)
        Total_Current_Debt= np.asfarray(TCD)

    for a in rowS3TA:
        TangAsst.append(a[0].value)
        Tangible_Assets = np.asfarray(TangAsst)

    Solvency_Ratio = (Total_Non_Current_Debt+Total_Current_Debt)/Tangible_Assets
    return (Solvency_Ratio)

def Propreitary_Ratio():
    SF=[]
    TA=[]
    Shareholders_Fund = []
    Total_Assets = []
    Propreitary_Ratio = []
    rowS1SF = sheet1.iter_cols(min_row=10, max_row=10, min_col=2, max_col=6)
    rowS1TA = sheet1.iter_cols(min_row=44, max_row=44, min_col=2, max_col=6)
    for a in rowS1SF:
        SF.append(a[0].value)
        Shareholders_Fund= np.asfarray(SF)
    for a in rowS1TA:
        TA.append(a[0].value)
        Total_Assets = np.asfarray(TA)

    rowS3SF = sheet3.iter_cols(min_row=10, max_row=10, min_col=2, max_col=6)
    rowS3TA = sheet3.iter_cols(min_row=44, max_row=44, min_col=2, max_col=6)
    for a in rowS3SF:
        SF.append(a[0].value)
        Shareholders_Fund= np.asfarray(SF)
    for a in rowS3TA:
        TA.append(a[0].value)
        Total_Assets = np.asfarray(TA)

    Propreitary_Ratio=Shareholders_Fund/Total_Assets
    return Propreitary_Ratio

def FixedAsset_Ratio():
    TFA=[]
    NCIN = []
    SF=[]
    TNCL = []
    Shareholders_Fund = []
    TotalFixed_Assets = []
    Total_Non_Current_Liab = []
    Non_Current_Investments = []
    FixedAsset_Ratio = []

    rowS1SF = sheet1.iter_cols(min_row=10, max_row=10, min_col=2, max_col=6)
    rowS1TFA = sheet1.iter_cols(min_row=30, max_row=30, min_col=2, max_col=6)
    rowS1TNCL = sheet1.iter_cols(min_row=12, max_row=12, min_col=2, max_col=6)
    rowS1NCIN = sheet1.iter_cols(min_row=12, max_row=12, min_col=2, max_col=6)

    for a in rowS1SF:
        SF.append(a[0].value)
        Shareholders_Fund= np.asfarray(SF)
    for a in rowS1TFA:
        TFA.append(a[0].value)
        Total_Assets = np.asfarray(TFA)
    for a in rowS1NCIN:
        NCIN.append(a[0].value)
        Non_Current_Investments= np.asfarray(NCIN)
    for a in rowS1TNCL:
        TNCL.append(a[0].value)
        Total_Non_Current_Liab = np.asfarray(TNCL)

    rowS3SF = sheet3.iter_cols(min_row=10, max_row=10, min_col=2, max_col=6)
    rowS3TFA = sheet3.iter_cols(min_row=30, max_row=30, min_col=2, max_col=6)
    rowS3TNCL = sheet3.iter_cols(min_row=12, max_row=12, min_col=2, max_col=6)
    rowS3NCIN = sheet3.iter_cols(min_row=12, max_row=12, min_col=2, max_col=6)

    for a in rowS3SF:
        SF.append(a[0].value)
        Shareholders_Fund= np.asfarray(SF)
    for a in rowS3TFA:
        TFA.append(a[0].value)
        Total_Assets = np.asfarray(TFA)
    for a in rowS3NCIN:
        NCIN.append(a[0].value)
        Non_Current_Investments= np.asfarray(NCIN)
    for a in rowS3TNCL:
        TNCL.append(a[0].value)
        Total_Non_Current_Liab = np.asfarray(TNCL)

    FixedAsset_Ratio=(Total_Assets+Non_Current_Investments)/(Shareholders_Fund+Total_Non_Current_Liab)
    return FixedAsset_Ratio


main()
