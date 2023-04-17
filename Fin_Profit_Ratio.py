import pandas as pd
import numpy as np
import openpyxl

workbook = openpyxl.load_workbook("Financial_Statement.xlsx")
sheet1 = workbook['balance sheet 2018-2022']
sheet2 = workbook['P&L 2018-2022']
sheet3 = workbook['balance sheet 2013-2017']
sheet4 = workbook['P&L 2013-2017']

def main():

    array1 = np.array(Net_Profit_Ratio())
    array2 = np.array(Gross_Profit_Ratio())
    array3 = np.array(Operating_Profit_Ratio())
    array4 = np.array(Asset_Turnover_Ratio())
    array5 = np.array(Return_On_Assets())
    array6 = np.array(Return_On_Equity())


    Header = ['Mar 22', 'Mar 21', 'Mar 20', 'Mar 19', 'Mar 18', 'Mar 17', 'Mar 16', 'Mar 15', 'Mar 14', 'Mar 13']
    df = pd.DataFrame ({' ' : Header,
                        'Net Profit Ratio': array1,
                        'Gross Profit Ratio': array2,
                        'Operating Profit Ratio': array3,
                        'Asset Turnover Ratio': array4,
                        'Return On Assets': array5,
                        'Return On Equity': array6})
    with pd.ExcelWriter(f'Financial_Ratio.xlsx', mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name = 'Profitability Ratios')





def Net_Profit_Ratio():
    TP=[]
    TOR=[]
    Total_Profit = []
    Total_Operating_Revenue = []
    Net_Profit_Ratio = []
    rowS2TP = sheet2.iter_cols(min_row=32, max_row=32, min_col=2, max_col=6)
    rowS2OR = sheet2.iter_cols(min_row=8, max_row=8, min_col=2, max_col=6)
    for a in rowS2TP:
        TP.append(a[0].value)
        Total_Profit= np.array(list(map(np.float_, TP)))
    for a in rowS2OR:
        TOR.append(a[0].value)
        Total_Operating_Revenue = np.array(list(map(np.float_, TOR)))

    rowS4TP = sheet4.iter_cols(min_row=32, max_row=32, min_col=2, max_col=6)
    rowS4OR = sheet4.iter_cols(min_row=8, max_row=8, min_col=2, max_col=6)
    for a in rowS4TP:
        TP.append(a[0].value)
        Total_Profit= np.array(list(map(np.float_, TP)))
    for a in rowS4OR:
        TOR.append(a[0].value)
        Total_Operating_Revenue = np.array(list(map(np.float_, TOR)))

    Net_Profit_Ratio=Total_Profit/Total_Operating_Revenue
    return(Net_Profit_Ratio)

def Gross_Profit_Ratio():
    COMC = []
    POSIT = []
    CINV= []
    TOR = []
    CostOfMat = []
    StockInTrade = []
    ChangeInvent = []
    Total_Op_Revenue = []
    CostOfGoodsSold = []
    Gross_Profit_Ratio = []

    rowS2COMC = sheet2.iter_cols(min_row=12, max_row=12, min_col=2, max_col=6)
    rowS2POSIT = sheet2.iter_cols(min_row=13, max_row=13, min_col=2, max_col=6)
    rowS2CINV = sheet2.iter_cols(min_row=14, max_row=14, min_col=2, max_col=6)
    rowS2OR = sheet2.iter_cols(min_row=8, max_row=8, min_col=2, max_col=6)

    for a in rowS2COMC:
        COMC.append(a[0].value)
        CostOfMat= np.array(list(map(np.float_, COMC)))
    for a in rowS2POSIT:
        POSIT.append(a[0].value)
        StockInTrade= np.array(list(map(np.float_, POSIT)))
    for a in rowS2CINV:
        CINV.append(a[0].value)
        ChangeInvent = np.array(list(map(np.float_, CINV)))
    for a in rowS2OR:
        TOR.append(a[0].value)
        Total_Op_Revenue = np.array(list(map(np.float_, TOR)))

    rowS4COMC = sheet4.iter_cols(min_row=12, max_row=12, min_col=2, max_col=6)
    rowS4POSIT = sheet4.iter_cols(min_row=13, max_row=13, min_col=2, max_col=6)
    rowS4CINV = sheet4.iter_cols(min_row=14, max_row=14, min_col=2, max_col=6)
    rowS4OR = sheet4.iter_cols(min_row=8, max_row=8, min_col=2, max_col=6)

    for a in rowS4COMC:
        COMC.append(a[0].value)
        CostOfMat= np.array(list(map(np.float_, COMC)))
    for a in rowS4POSIT:
        POSIT.append(a[0].value)
        StockInTrade= np.array(list(map(np.float_, POSIT)))
    for a in rowS4CINV:
        CINV.append(a[0].value)
        ChangeInvent = np.array(list(map(np.float_, CINV)))
    for a in rowS4OR:
        TOR.append(a[0].value)
        Total_Op_Revenue = np.array(list(map(np.float_, TOR)))


    CostOfGoodsSold = CostOfMat+StockInTrade+ChangeInvent
    Gross_Profit_Ratio = (Total_Op_Revenue-CostOfGoodsSold)/Total_Op_Revenue
    return(Gross_Profit_Ratio)

def Operating_Profit_Ratio():
    TOE = []
    FC = []
    TOR= []
    Total_Op_Expense = []
    Finance_Cost = []
    Total_Op_Revenue = []
    EBIT = []
    Operating_Profit_Ratio = []

    rowS2TOE = sheet2.iter_cols(min_row=20, max_row=20, min_col=2, max_col=6)
    rowS2FC = sheet2.iter_cols(min_row=17, max_row=17, min_col=2, max_col=6)
    rowS2OR = sheet2.iter_cols(min_row=8, max_row=8, min_col=2, max_col=6)

    for a in rowS2TOE:
        TOE.append(a[0].value)
        Total_Op_Expense = np.array(list(map(np.float_, TOE)))
    for a in rowS2FC:
        FC.append(a[0].value)
        Finance_Cost = np.array(list(map(np.float_, FC)))
    for a in rowS2OR:
        TOR.append(a[0].value)
        Total_Op_Revenue = np.array(list(map(np.float_, TOR)))

    rowS4TOE = sheet4.iter_cols(min_row=20, max_row=20, min_col=2, max_col=6)
    rowS4FC = sheet4.iter_cols(min_row=17, max_row=17, min_col=2, max_col=6)
    rowS4OR = sheet4.iter_cols(min_row=8, max_row=8, min_col=2, max_col=6)

    for a in rowS4TOE:
        TOE.append(a[0].value)
        Total_Op_Expense = np.array(list(map(np.float_, TOE)))
    for a in rowS4FC:
        FC.append(a[0].value)
        Finance_Cost = np.array(list(map(np.float_, FC)))
    for a in rowS4OR:
        TOR.append(a[0].value)
        Total_Op_Revenue = np.array(list(map(np.float_, TOR)))

    EBIT = Total_Op_Revenue-(Total_Op_Expense-Finance_Cost)
    Operating_Profit_Ratio = EBIT/Total_Op_Revenue
    return(Operating_Profit_Ratio)


def Asset_Turnover_Ratio():
    TA=[]
    TOR=[]
    Total_Asset = []
    Total_Operating_Revenue = []
    Average_Asset = []
    Asset_Turnover_Ratio = []

    rowS1TA = sheet1.iter_cols(min_row=44, max_row=44, min_col=2, max_col=6)
    rowS2OR = sheet2.iter_cols(min_row=8, max_row=8, min_col=2, max_col=6)
    for a in rowS1TA:
        TA.append(a[0].value)
        Total_Asset= np.array(list(map(np.float_, TA)))
    for a in rowS2OR:
        TOR.append(a[0].value)
        Total_Operating_Revenue = np.array(list(map(np.float_, TOR)))

    rowS3TA = sheet3.iter_cols(min_row=44, max_row=44, min_col=2, max_col=6)
    rowS4OR = sheet4.iter_cols(min_row=8, max_row=8, min_col=2, max_col=6)
    for a in rowS3TA:
        TA.append(a[0].value)
        Total_Asset= np.array(list(map(np.float_, TA)))
    for a in rowS4OR:
        TOR.append(a[0].value)
        Total_Operating_Revenue = np.array(list(map(np.float_, TOR)))

    for i in range(len(Total_Asset)-1):
        result = (Total_Asset[i] + Total_Asset[i+1])/2
        Average_Asset.append(result)

    last_value = Total_Asset[-1]
    Average_Asset.append(last_value)

    Asset_Turnover_Ratio = (Total_Operating_Revenue/Average_Asset)
    return(Asset_Turnover_Ratio)

def Return_On_Assets():
    TA=[]
    PAT=[]
    Total_Asset = []
    Profit_After_Tax = []
    Average_Asset = []
    Return_On_Assets = []

    rowS1TA = sheet1.iter_cols(min_row=44, max_row=44, min_col=2, max_col=6)
    rowS2PAT = sheet2.iter_cols(min_row=8, max_row=8, min_col=2, max_col=6)
    for a in rowS1TA:
        TA.append(a[0].value)
        Total_Asset= np.array(list(map(np.float_, TA)))
    for a in rowS2PAT:
        PAT.append(a[0].value)
        Profit_After_Tax = np.array(list(map(np.float_, PAT)))

    rowS3TA = sheet3.iter_cols(min_row=44, max_row=44, min_col=2, max_col=6)
    rowS4PAT = sheet4.iter_cols(min_row=32, max_row=32, min_col=2, max_col=6)
    for a in rowS3TA:
        TA.append(a[0].value)
        Total_Asset= np.array(list(map(np.float_, TA)))
    for a in rowS4PAT:
        PAT.append(a[0].value)
        Profit_After_Tax = np.array(list(map(np.float_, PAT)))

    for i in range(len(Total_Asset)-1):
        result = (Total_Asset[i] + Total_Asset[i+1])/2
        Average_Asset.append(result)

    last_value = Total_Asset[-1]
    Average_Asset.append(last_value)

    Return_On_Assets = (Profit_After_Tax/Average_Asset)
    return(Return_On_Assets)

def Return_On_Equity():
    TSE=[]
    PAT=[]
    Total_Shareholder_Equity = []
    Profit_After_Tax = []
    Return_On_Equity = []

    rowS1TSE = sheet1.iter_cols(min_row=10, max_row=10, min_col=2, max_col=6)
    rowS2PAT = sheet2.iter_cols(min_row=32, max_row=32, min_col=2, max_col=6)
    for a in rowS1TSE:
        TSE.append(a[0].value)
        Total_Shareholder_Equity= np.array(list(map(np.float_, TSE)))
    for a in rowS2PAT:
        PAT.append(a[0].value)
        Profit_After_Tax = np.array(list(map(np.float_, PAT)))

    rowS3TSE = sheet3.iter_cols(min_row=10, max_row=10, min_col=2, max_col=6)
    rowS4PAT = sheet4.iter_cols(min_row=32, max_row=32, min_col=2, max_col=6)
    for a in rowS3TSE:
        TSE.append(a[0].value)
        Total_Shareholder_Equity= np.array(list(map(np.float_, TSE)))
    for a in rowS4PAT:
        PAT.append(a[0].value)
        Profit_After_Tax = np.array(list(map(np.float_, PAT)))

    Return_On_Equity = Profit_After_Tax/Total_Shareholder_Equity
    return(Return_On_Equity)






main()
