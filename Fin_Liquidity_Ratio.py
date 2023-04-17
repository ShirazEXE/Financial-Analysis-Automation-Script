import openpyxl
import pandas as pd
import numpy as np


workbook = openpyxl.load_workbook("Financial_Statement.xlsx")
sheet1 = workbook['balance sheet 2018-2022']
sheet2 = workbook['P&L 2018-2022']
sheet3 = workbook['balance sheet 2013-2017']
sheet4 = workbook['P&L 2013-2017']

def main():

    array1 = np.array(CurrentRatio())
    array2 = np.array(QuickRatio())
    array3 = np.array(AcidRatio())


    Header = ['Mar 22', 'Mar 21', 'Mar 20', 'Mar 19', 'Mar 18', 'Mar 17', 'Mar 16', 'Mar 15', 'Mar 14', 'Mar 13']
    df = pd.DataFrame ({' ' : Header,
                        'Current Ratio': array1,
                        'Quick Ratio': array2,
                        'Acid Test Ratio': array3})

    with pd.ExcelWriter(f'Financial_Ratio.xlsx', mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name = 'Liquidity Ratios')

def CurrentRatio():
    CA=[]
    CL=[]
    CurAsst = []
    CurLib = []
    CurRatio = []
    rowS1CA = sheet1.iter_cols(min_row=43, max_row=43, min_col=2, max_col=6)
    rowS1CL = sheet1.iter_cols(min_row=22, max_row=22, min_col=2, max_col=6)
    for a in rowS1CA:
        CA.append(a[0].value)
        CurAsst= np.array(list(map(np.float_, CA)))
    for a in rowS1CL:
        CL.append(a[0].value)
        CurLib = np.array(list(map(np.float_, CL)))

    rowS3CA = sheet3.iter_cols(min_row=43, max_row=43, min_col=2, max_col=6)
    rowS3CL = sheet3.iter_cols(min_row=22, max_row=22, min_col=2, max_col=6)
    for a in rowS3CA:
        CA.append(a[0].value)
        CurAsst= np.array(list(map(np.float_, CA)))
    for a in rowS3CL:
        CL.append(a[0].value)
        CurLib = np.array(list(map(np.float_, CL)))

    CurRatio = CurAsst/CurLib
    return CurRatio


def QuickRatio():
    CA = []
    INV = []
    CL = []
    CurAsst = []
    Inventory = []
    CurLib = []
    QuickRatio = []

    rowS1CA = sheet1.iter_cols(min_row=43, max_row=43, min_col=2, max_col=6)
    rowS1INV = sheet1.iter_cols(min_row=38, max_row=38, min_col=2, max_col=6)
    rowS1CL = sheet1.iter_cols(min_row=22, max_row=22, min_col=2, max_col=6)
    for a in rowS1CA:
        CA.append(a[0].value)
        CurAsst= np.array(list(map(np.float_, CA)))
    for a in rowS1INV:
        INV.append(a[0].value)
        Inventory=np.array(list(map(np.float_, INV)))
    for a in rowS1CL:
        CL.append(a[0].value)
        CurLib = np.array(list(map(np.float_, CL)))

    rowS3CA = sheet3.iter_cols(min_row=43, max_row=43, min_col=2, max_col=6)
    rowS3INV = sheet3.iter_cols(min_row=38, max_row=38, min_col=2, max_col=6)
    rowS3CL = sheet3.iter_cols(min_row=22, max_row=22, min_col=2, max_col=6)
    for a in rowS3CA:
        CA.append(a[0].value)
        CurAsst= np.array(list(map(np.float_, CA)))
    for a in rowS3INV:
        INV.append(a[0].value)
        Inventory=np.array(list(map(np.float_, INV)))
    for a in rowS3CL:
        CL.append(a[0].value)
        CurLib = np.array(list(map(np.float_, CL)))

    QuickRatio = (CurAsst-Inventory)/CurLib
    return QuickRatio


def AcidRatio():
    CASH = []
    AR = []
    CL = []
    CashAsst = []
    AccRecAsst = []
    CurLib = []
    QuickRatio = []

    rowS1CASH = sheet1.iter_cols(min_row=40, max_row=40, min_col=2, max_col=6)
    rowS1AR = sheet1.iter_cols(min_row=39, max_row=39, min_col=2, max_col=6)
    rowS1CL = sheet1.iter_cols(min_row=22, max_row=22, min_col=2, max_col=6)
    for a in rowS1CASH:
        CASH.append(a[0].value)
        CashAsst= np.array(list(map(np.float_, CASH)))
    for a in rowS1AR:
        AR.append(a[0].value)
        AccRecAsst=np.array(list(map(np.float_, AR)))
    for a in rowS1CL:
        CL.append(a[0].value)
        CurLib = np.array(list(map(np.float_, CL)))

    rowS3CASH = sheet3.iter_cols(min_row=40, max_row=40, min_col=2, max_col=6)
    rowS3AR = sheet3.iter_cols(min_row=39, max_row=39, min_col=2, max_col=6)
    rowS3CL = sheet3.iter_cols(min_row=22, max_row=22, min_col=2, max_col=6)
    for a in rowS3CASH:
        CASH.append(a[0].value)
        CashAsst= np.array(list(map(np.float_, CASH)))
    for a in rowS3AR:
        AR.append(a[0].value)
        AccRecAsst=np.array(list(map(np.float_, AR)))
    for a in rowS3CL:
        CL.append(a[0].value)
        CurLib = np.array(list(map(np.float_, CL)))

    AcidRatio=(CashAsst+AccRecAsst)/CurLib
    return AcidRatio

main()
